import random
import openpyxl as xl

path = 'C:/Users/Harry/Documents/CSWorkspaces/pokemon-predictor/pokemon.xlsx'

def loadWorkbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb.active

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 1)
        ageValue = random.randint(10,80)
        ageValueCell = cell
        ageValueCell.value = ageValue

        cell = sheet.cell(row, 2)
        genderValue = random.randint(0,1) # 0 is women, 1 is men
        genderValueCell = cell
        genderValueCell.value = genderValue
        

    wb.save('C:/Users/Harry/Documents/CSWorkspaces/pokemon-predictor/pokemon.xlsx')

loadWorkbook(path)